from urllib.error import HTTPError
import time
import urllib
from urllib import request
from bs4 import BeautifulSoup
import re
import requests
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import socket
import json
import re
import os
import tempfile
import csv
import pandas as pd
baseUrl = "https://www.programmableweb.com"
# 以上这些灰色的导入虽然在这个文件没用到，但是引用他的文件用到了这些文件
# 这里是一些公共的方法
# 检测是否需要重发，网络不稳定是进行重发
# def isNeedReSendUrl(BS, url):
#     # 设置一个重发的次数，发送无效的url地址时陷入死循环
#     count = 0
#     # 休眠的初始值为10
#     second = 10
#     while BS is None:
#         # 设置一个休眠时间，等一会儿再发
#         print("休眠" + str(second) + "后重发:" + url)
#         time.sleep(second)
#         BS = getHtml(url)
#         second = second + 5
#         count = count + 1
#         # 这是是避免一些无效的url地址，但是执行到了break基本上代表着代码有问题或者网络长时间没有链接上
#         if count == 10:
#             break
#     return BS
#
# 获取html代码,这个是本机的ip来发送
def getHtml(url):
    try:
        response = urllib.request.urlopen(url, timeout=50)
        html = response.read().decode("utf-8")
        BS = BeautifulSoup(html, "html.parser")
        return BS
    except Exception as e:
        print("获取html失败，请先检查网络是否通畅")
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
        return None
#注意，这个的proxy的ip需要经常更换，为什么呢，因为这里面的ip的时长只能限制在1-3分钟之内，超过这个时间就可能失效
def get_proxy():
    try:
        proxy_url = 'http://http2.9vps.com/getip.asp?username=1376493805&pwd=751a76e73aa3b7f6ae8d66c588774555&geshi=1&fenge=1&fengefu=&getnum=1'
        # headers = {
        #     'User-Agent': 'Mozilla/4.0 (compatible; MSIE 9.0; Windows NT 6.1)'
        # }
        ip = requests.get(proxy_url).text
        proxy_host = ip.strip()
        proxy = {

            # 'http': proxy_host,
            'https': proxy_host,
        }
        # print(proxy)
        return proxy
    except:
        hostname = socket.gethostname()
        ip = socket.gethostbyname(hostname)
        return {
            'https': ip,
        }
proxy = get_proxy()
#本机ip重发失败的次数
failCount = 0
def isNeedReSendUrl(BS, url):
    global proxy
    global failCount
    # 设置一个重发的次数，发送无效的url地址时陷入死循环
    count = 1
    # 休眠的初始值为10
    second = 1
    # proxyCount = 0  # 统计切换的代理ip数量
    # count2 = 0  # 统计代理重发次数
    while BS is None:
        # 设置一个休眠时间，等一会儿再发
        # print("休眠" + str(second) + "后重发:" + url)
        print("第"+str(count)+"次尝试重发", "先休眠"+str(second)+"s")
        time.sleep(second)
        second = second + 2
        # 可能爬虫爬久啦，导致本机的ip被封啦，这时候需要切换使用代理ip,相同的的请求超过10次,也可以切换使用代理
        # if count <= 10 & failCount <= 2000:
        # failCount = failCount + 1
        BS = getHtml(url)
        # 由于免费的代理不好用，所有先暂时不考虑使用代理

        # else:
        #     BS = getHtml2(url)
        #     count2 = count2 + 1
        #     # 这是是避免一些无效的url地址，但是执行到了break基本上代表着代码有问题或者网络长时间没有链接上，每个代理ip尝试发送5次
        #     if count2 == 6:
        #         proxy = get_proxy()
        #         print(proxy)
        #         proxyCount = proxyCount + 1
        #         count2 = 0
        #     # 换了100个代理ip依旧失败,那就结束循环吧
        #     if proxyCount == 100:
        #         break
        #     # 由于代理不好用，总是请求不到数据，所以使用一段时间后，可能自己本机的ip解封啦，所以需要切换使用本机的ip请求数据
        #     if failCount > 2000:
        #         failCount = failCount + 1
        #     if failCount >= 3000:
        #         failCount = 0
        count = count + 1
        if count > 9:
            return None

    return BS
# 获取html代码,使用代理发送
def getHtml2(url):

    try:
        handler = urllib.request.ProxyHandler(proxy)
        opener = urllib.request.build_opener(handler)
        req = urllib.request.Request(url)
        response = opener.open(req, timeout=50)
        # response = urllib.request.urlopen(url, timeout=50)
        html = response.read().decode("utf-8")
        BS = BeautifulSoup(html, "html.parser")
        return BS
    except Exception as e:
        print("获取html失败，请先检查网络是否通畅")
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
        return None
# 数组元素的连接
def partitionedArray(array):
    return "###".join(array)

# 去掉amp;
def subStringAmp(value):
    return str(value).replace("amp;", "")

#把数据保存在excel表格中
# row = 0
# wb = Workbook()
# ws = wb.active
# ws.title = u'API数据'
# #按行存储字典数据
# def saveDataInExcel(singleData, wb, ws, row, fileName):
#     col = 0
#     for key in singleData.keys():
#         content = ILLEGAL_CHARACTERS_RE.sub(r'', singleData[key])
#         if row == 0:
#             key = ILLEGAL_CHARACTERS_RE.sub(r'', key)
#             ws.cell(row + 1, col + 1).value = key
#             ws.cell(row + 2, col + 1).value = content
#             col = col + 1
#         else:
#             ws.cell(row + 2, col + 1).value = content
#             col = col + 1
#     print("row:", row + 2)
#     wb.save('./data/'+fileName+'.xlsx')
# 按行存储字典数据，保存数据到excel
def saveDataInExcel(singleData, row, fileName, sheetName):
    try:
        filePath = './data/'+fileName+'.xlsx'
        wb = load_workbook(filePath)
        ws = wb[sheetName]
    except Exception:
        print("没有这个表格,先进行创建一个")
        wb = Workbook()
        ws = wb.active
        ws.title = sheetName
    finally:
        col = 0
        for key in singleData.keys():
            content = ILLEGAL_CHARACTERS_RE.sub(r'', singleData[key])
            if row == 0:
                key = ILLEGAL_CHARACTERS_RE.sub(r'', key)
                ws.cell(row + 1, col + 1).value = key
                ws.cell(row + 2, col + 1).value = content
                col = col + 1
            else:
                ws.cell(row + 2, col + 1).value = content
                col = col + 1
        print("row:", row + 2)
        wb.save('./data/'+fileName+'.xlsx')

# 读取excel表格数据，然后把excel转换成csv
def excelToCSV(fileName):
    data = pd.read_excel('./data/'+fileName+'.xlsx', fileName, index_col=0)
    data.to_csv('./data/'+fileName+'.csv', encoding='utf-8')
# 按行存储字典数据,保存数据到excel
# def saveDataInCSV(singleData, fileName, row):
#     data = []
#     header = list(singleData.keys())
#     data.append(singleData)
#     with open('./data/'+fileName+'.csv', 'a', newline='', encoding='utf-8') as f:
#         writer = csv.DictWriter(f, fieldnames=header)  # 提前预览列名，当下面代码写入数据时，会将其一一对应。
#         if row == 0:
#             writer.writeheader()  # 写入列名
#             writer.writerows(data)  # 写入数据
#         else:
#             writer.writerows(data)  # 写入数据
#     print("row:", row + 2)

def savePositionInfo(preName, page, count):
    properties = getProperties()
    # properties.put(preName+'.url', apiUrl)
    properties.put(preName+'.page', page)
    properties.put(preName + '.count', count)
    print("保存位置信息")
    # file_name = './data/'+fileName+'PositionInfo.txt'
    # data = {
    #     'name': name,
    #     'apiUrl': apiUrl,
    #     'page': page,
    #     'count': count
    # }
    # with open(file_name, 'w') as file_obj:
    #     json.dump(data, file_obj)
    #     print("把当前的位置写入了文件")


def getProperties():
    properties_path = "./data/field.properties"
    properties = Properties(properties_path)
    return properties

def initProperties():
    properties = getProperties()
    properties.put("flag", "false")
    # apis模块回到最初的位置
    properties.put("apis.name", "apis")
    properties.put("apis.flag", "false")
    properties.put("apis.page", str(0))
    properties.put("apis.count", str(0))
    # mashups模块回到最初的位置
    properties.put("mashups.name", "mashups")
    properties.put("mashups.flag", "false")
    properties.put("mashups.page", str(0))
    properties.put("mashups.count", str(0))
    # sampleSourceCode模块回到最初的位置
    properties.put("sampleSourceCode.name", "sampleSourceCode")
    properties.put("sampleSourceCode.flag", "false")
    properties.put("sampleSourceCode.page", str(0))
    properties.put("sampleSourceCode.count", str(0))
    # sdks模块回到最初的位置
    properties.put("sdks.name", "sdks")
    properties.put("sdks.flag", "false")
    properties.put("sdks.page", str(0))
    properties.put("sdks.count", str(0))


class Properties:
    def __init__(self, file_name):
        self.file_name = file_name
        self.properties = {}
        try:
            fopen = open(self.file_name, 'r')
            for line in fopen:
                line = line.strip()
                if line.find('=') > 0 and not line.startswith('#'):
                    strs = line.split('=')
                    self.properties[strs[0].strip()] = strs[1].strip()
        except Exception:
            raise
        else:
            fopen.close()

    def has_key(self, key):
        return key in self.properties

    def get(self, key, default_value=''):
        if key in self.properties:
            return self.properties[key]
        return default_value

    def put(self, key, value):
        self.properties[key] = value
        replace_property(self.file_name, key + '=.*', key + '=' + value, True)


def replace_property(file_name, from_regex, to_str, append_on_not_exists=True):
    tmpfile = tempfile.TemporaryFile()

    if os.path.exists(file_name):
        r_open = open(file_name, 'r')
        pattern = re.compile(r'' + from_regex)
        found = None
        for line in r_open:
            if pattern.search(line) and not line.strip().startswith('#'):
                found = True
                line = re.sub(from_regex, to_str, line)
            tmpfile.write(line.encode())
        if not found and append_on_not_exists:
            tmpfile.write(('\n' + to_str).encode())
        r_open.close()
        tmpfile.seek(0)

        content = tmpfile.read()

        if os.path.exists(file_name):
            os.remove(file_name)

        w_open = open(file_name, 'wb')
        w_open.write(content)
        w_open.close()

        tmpfile.close()
    else:
        print("file %s not found" % file_name)


